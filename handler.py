from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from states import user_state
from config import ADMIN_ID, SUPPORT_USERNAME, BOT_NAME, FORCE_JOIN_CHANNEL, FORCE_JOIN_LINK
from buttons import main_menu, shop_menu, deposit_menu, product_menu, quantity_menu, force_join_menu
from admin import admin_buttons
from database import create_user, get_balance, update_balance, add_order, get_orders, get_order_by_id, update_order_status, c, get_all_users, get_stock_count, take_codes, add_stock, add_referral, activate_referral_bonus, get_refer_stats, get_all_stock
from bot import bot
import io

MAINTENANCE_MODE = False
MAINT_MSG = "🔧 Bot Update চলছে... ⏳\n\n📢 Update Complete হলে Bot/Channel-এ জানিয়ে দেওয়া হবে।\n\n🙏 সবাই একটু অপেক্ষা করুন। ❤"

def is_maintenance_block(user_id):
    if MAINTENANCE_MODE and user_id != ADMIN_ID:
        return True
    return False

def is_user_joined(bot, user_id):
    try:
        member = bot.get_chat_member(FORCE_JOIN_CHANNEL, user_id)
        return member.status in ['member', 'administrator', 'creator']
    except Exception as e:
        return False

def format_proxy_message(name, qty, total_price, codes, is_manual=False):
    if is_manual:
        full_text = f"✅ DELIVERY SUCCESSFUL!\n───────────────\n\n"
    else:
        full_text = f"✅ DELIVERY SUCCESSFUL!\n───────────────\n\n📦 Product: {name}\n🔢 Quantity: {qty} pcs\n💰 Total: {total_price} BDT\n\n"
    for idx, code in enumerate(codes, 1):
        try:
            c_parts = code.strip().split(":")
            if len(c_parts) >= 4:
                host = c_parts[0]
                port = c_parts[1]
                user = c_parts[2]
                pwd = ":".join(c_parts[3:])
            else:
                host = code
                port = user = pwd = "N/A"
        except:
            host = code
            port = user = pwd = "N/A"
        if len(codes) > 1:
            full_text += f"**Proxy {idx}:**\n"
        full_text += f"IP: `{host}`\n"
        full_text += f"🔌 Port: `{port}`\n"
        full_text += f"👤 User: `{user}`\n"
        full_text += f"🔑 Pass: `{pwd}`\n\n"
    full_text += f"✅ আপনার অর্ডারটি কমপ্লিট হয়েছে!\nআমাদের উপর ভরসা রাখার জন্য ধন্যবাদ। ❤"
    return full_text

def register_handlers(bot):

    @bot.message_handler(commands=["maintenance"])
    def maintenance_toggle(message):
        global MAINTENANCE_MODE
        if message.from_user.id != ADMIN_ID:
            return
        args = message.text.split()
        if len(args) < 2:
            status = "ON 🟢" if MAINTENANCE_MODE else "OFF 🔴"
            bot.reply_to(message, f"🛠️ Maintenance Mode: {status}\n\n/maintenance on / off")
            return
        if args[1].lower() == "on":
            MAINTENANCE_MODE = True
            bot.reply_to(message, "✅ Maintenance ON!")
        elif args[1].lower() == "off":
            MAINTENANCE_MODE = False
            bot.reply_to(message, "✅ Maintenance OFF!")

    @bot.message_handler(commands=["start"])
    def start(message):
        user_id = message.from_user.id
        if is_maintenance_block(user_id):
            bot.reply_to(message, MAINT_MSG)
            return
        args = message.text.split()
        if len(args) > 1:
            try:
                ref_id = int(args[1])
                if ref_id != user_id:
                    create_user(user_id, ref_id)
                    add_referral(ref_id, user_id)
            except: create_user(user_id)
        else:
            create_user(user_id)
        if user_id != ADMIN_ID and not is_user_joined(bot, user_id):
            bot.send_message(message.chat.id, f"⚠ Bot use korte hole amader channel e join korte hobe!\n\n📢 Channel: {FORCE_JOIN_CHANNEL}\n\nJoin kore Verify koro.", reply_markup=force_join_menu())
            return
        bot.send_message(message.chat.id, f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", reply_markup=main_menu())

    @bot.message_handler(commands=["admin"])
    def admin(message):
        if message.from_user.id != ADMIN_ID:
            bot.reply_to(message, "❌ Access Denied")
            return
        bot.send_message(message.chat.id, "👑 Admin Panel", reply_markup=admin_buttons())

    @bot.callback_query_handler(func=lambda call: True)
    def callback(call):
        msg_id = call.message.message_id
        chat_id = call.message.chat.id
        user_id = call.from_user.id
        if is_maintenance_block(user_id):
            try: bot.answer_callback_query(call.id, "🔧 Bot Update চলছে... ⏳", show_alert=True)
            except: pass
            try: bot.send_message(chat_id, MAINT_MSG)
            except: pass
            return
        create_user(user_id)
        if call.data == "verify_join":
            if is_user_joined(bot, user_id):
                try: bot.edit_message_text(f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                except: pass
            else:
                bot.answer_callback_query(call.id, "❌ Tumi ekhono Channel e Join koro nai! Age Join koro.", show_alert=True)
            return
        if user_id != ADMIN_ID and not is_user_joined(bot, user_id):
            try: bot.answer_callback_query(call.id, "⚠ Age Channel Join Koro!")
            except: pass
            bot.send_message(chat_id, f"⚠ Age Channel Join Koro!\n{FORCE_JOIN_CHANNEL}", reply_markup=force_join_menu())
            return
        if call.data == "shop":
            try: bot.edit_message_text("🛒 Select Category", chat_id=chat_id, message_id=msg_id, reply_markup=shop_menu())
            except: pass
        elif call.data == "vpn_list":
            try: bot.edit_message_text("🌐 VPN Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("vpn"))
            except: pass
        elif call.data == "proxy_list":
            try: bot.edit_message_text("🌍 Proxy Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("proxy"))
            except: pass
        elif call.data == "gmail_list":
            try: bot.edit_message_text("📧 Gmail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("gmail"))
            except: pass
        elif call.data == "outlook_list":
            try: bot.edit_message_text("📮 Outlook Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("outlook"))
            except: pass
        elif call.data == "hotmail_list":
            try: bot.edit_message_text("📬 Hotmail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("hotmail"))
            except: pass
        elif call.data == "edumail_list":
            try: bot.edit_message_text("🎓 Edu Mail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("edumail"))
            except: pass
        elif call.data == "morelogin_list":
            try: bot.edit_message_text("🖥 Morelogin 100 Minutes", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("morelogin"))
            except: pass
        elif call.data == "noop":
            bot.answer_callback_query(call.id, "Quantity change korte + - use koro")
        elif call.data.startswith("select_qty|"):
            parts = call.data.split("|")
            category, name, price = parts[1], parts[2], float(parts[3])
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: 1", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, 1))
            except: pass
        elif call.data.startswith("qty_plus|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            qty += 1
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: {qty}", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, qty))
            except: pass
        elif call.data.startswith("qty_minus|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            if qty > 1: qty -= 1
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: {qty}", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, qty))
            except: pass
        elif call.data.startswith("custom_qty|"):
            parts = call.data.split("|")
            category, name, price = parts[1], parts[2], float(parts[3])
            user_state[user_id] = {"step": "custom_qty", "category": category, "name": name, "price": price}
            bot.send_message(chat_id, "📝 Koyta niba? Number likhe pathao")
        elif call.data.startswith("buy|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            total_price = price * qty
            balance = get_balance(user_id)
            if balance >= total_price:
                is_auto = False
                if category == "morelogin": is_auto = True
                elif category == "proxy" and "owl" in name.lower(): is_auto = True
                elif category in ["outlook", "edumail", "hotmail", "gmail"]: is_auto = True
                if is_auto:
                    available = get_stock_count(category, name)
                    if available < qty:
                        bot.send_message(ADMIN_ID, f"⚠ Stock sesh! {name} - {qty} pcs order asche but stock {available} pcs")
                        try: bot.edit_message_text(f"❌ Stock sesh! Available: {available} pcs", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                        except: pass
                        return
                update_balance(user_id, -total_price)
                if is_auto:
                    codes = take_codes(category, name, qty)
                    add_order(user_id, f"{name} x{qty}", total_price)
                    if qty > 5:
                        import openpyxl
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Delivery"
                        ws.append(["Product Name", "No", "Account Details"])
                        ws.column_dimensions['A'].width = 30
                        ws.column_dimensions['B'].width = 10
                        ws.column_dimensions['C'].width = 90
                        for i, code in enumerate(codes, 1): ws.append([name, i, code])
                        file_stream = io.BytesIO()
                        wb.save(file_stream)
                        file_stream.seek(0)
                        file_stream.name = f"{name.replace(' ','_')}_{qty}pcs.xlsx"
                        bot.send_document(user_id, file_stream, caption=f"✅ Order Delivered!\n\nProduct: {name}\nQuantity: {qty} pcs\nTotal: {total_price} BDT\n\nProblem hole {SUPPORT_USERNAME}")
                        try: bot.edit_message_text(f"✅ Order Complete! File diye disi", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                        except: pass
                    else:
                        is_proxy_product = (category == "proxy")
                        if is_proxy_product:
                            text = format_proxy_message(name, qty, total_price, codes)
                            bot.send_message(user_id, text, parse_mode="Markdown")
                        else:
                            text_codes = "\n".join([f"{i}. `{c}`" for i, c in enumerate(codes, 1)])
                            bot.send_message(user_id, f"✅ Order Delivered!\n\n📦 Product: {name}\n🔢 Quantity: {qty} pcs\n💰 Total: {total_price} BDT\n\n🔑 Codes:\n{text_codes}\n\nProblem hole {SUPPORT_USERNAME}", parse_mode="Markdown")
                        try: bot.edit_message_text(f"✅ Order Complete! Code text akare diye disi", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                        except: pass
                else:
                    oid = add_order(user_id, f"{name} x{qty}", total_price)
                    update_order_status(oid, "Pending")
                    try: bot.edit_message_text(f"✅ Order Confirmed!\n\nProduct: {name}\nQuantity: {qty} pcs\nTotal: {total_price} BDT\n\nAdmin 5-10 min er moddhe code diye dibe", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                    except: pass
                    bot.send_message(ADMIN_ID, f"🛒 New Manual Order\nOrder ID: {oid}\nUser: {user_id}\nProduct: {name} x{qty}\nTotal: {total_price} BDT\n\n/admin > Pending Orders")
            else:
                try: bot.edit_message_text(f"❌ Not Enough Balance\nYour Balance: {balance} BDT\nRequired: {total_price} BDT", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu())
                except: pass
        elif call.data == "deposit":
            try:
                bot.edit_message_text("💰 <b>ব্যালেন্স ডিপোজিট করুন</b>\n\nআপনার পছন্দের পেমেন্ট মেথডটি নিচ থেকে সিলেক্ট করুন।", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="HTML")
            except: pass
        elif call.data == "bkash":
            try:
                bot.edit_message_text("💳 <b>bKash Personal Payment</b>\n\nনাম্বার: <code>01603940061</code>\n\n🔹 <b>নির্দেশনা:</b>\n1. উপরের নাম্বারটি কপি করুন\n2. bKash App থেকে <b>Send Money</b> করুন\n3. পেমেন্ট সম্পন্ন হলে <b>Submit Payment</b> বাটনে ক্লিক করুন", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="HTML")
            except: pass
        elif call.data == "nagad":
            try:
                bot.edit_message_text("💳 <b>Nagad Personal Payment</b>\n\nনাম্বার: <code>01603940061</code>\n\n🔹 <b>নির্দেশনা:</b>\n1. উপরের নাম্বারটি কপি করুন\n2. Nagad App থেকে <b>Send Money</b> করুন\n3. পেমেন্ট সম্পন্ন হলে <b>Submit Payment</b> বাটনে ক্লিক করুন", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="HTML")
            except: pass
        elif call.data == "rocket":
            try:
                bot.edit_message_text("💳 <b>Rocket Personal Payment</b>\n\n🚫 <b>বর্তমানে বন্ধ আছে</b>\n\nঅনুগ্রহ করে bKash / Nagad / USDT এর মাধ্যমে ডিপোজিট করুন।", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="HTML")
            except: pass
        elif call.data == "usdt":
            try:
                bot.edit_message_text(
                    "💲 <b>USDT Payment</b>\n\n"
                    "🔹 <b>TRC20 (USDT):</b>\n"
                    "<code>TVRvRX3BZ9mrzQJgjTCryiyVChWmGZ9oJz</code>\n\n"
                    "🔹 <b>BEP20 (USDT):</b>\n"
                    "<code>0x0Bc20843c4452C6fAcAf7E1b757a00c0F79D6268</code>\n\n"
                    "🔹 <b>BINANCE ID:</b>\n"
                    "<code>910523531</code>\n\n"
                    "🔹 <b>নির্দেশনা:</b>\n"
                    "1. নেটওয়ার্ক অনুযায়ী এড্রেস কপি করুন\n"
                    "2. সঠিক নেটওয়ার্কে USDT পাঠান\n"
                    "3. পেমেন্ট সম্পন্ন হলে <b>Submit Payment</b> বাটনে ক্লিক করুন",
                    chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="HTML"
                )
            except: pass
        elif call.data == "submit_payment":
            user_state[user_id] = {"step": "amount"}
            bot.send_message(chat_id, "💰 Enter Deposit Amount")
        elif call.data == "admin_add_stock":
            if user_id != ADMIN_ID: return
            bot.send_message(chat_id, "📦 File pathao\n.txt ba.xlsx")
            user_state[user_id] = {"step": "wait_txt_file"}
        elif call.data == "admin_broadcast":
            if user_id != ADMIN_ID: return
            bot.send_message(chat_id, "📢 Broadcast message likhe pathao.")
            user_state[user_id] = {"step": "broadcast_msg"}
        elif call.data == "admin_add_balance":
            if user_id != ADMIN_ID: return
            bot.send_message(chat_id, "👤 User er Telegram ID dao")
            user_state[user_id] = {"step": "admin_user_id"}
        elif call.data == "admin_orders":
            if user_id != ADMIN_ID: return
            try:
                c.execute("SELECT id, user_id, product, price, status FROM orders ORDER BY id DESC LIMIT 20")
                orders = c.fetchall()
                text = "📦 No orders yet" if not orders else "📦 Last 20 Orders\n"+"\n".join([f"ID: {x[0]} | User: {x[1]}\nProduct: {x[2]}\nPrice: {x[3]} BDT | {x[4]}\n" for x in orders])
                bot.send_message(chat_id, text)
            except Exception as e:
                from database import conn; conn.rollback()
                bot.send_message(chat_id, f"Error: {e}")
        elif call.data == "admin_stock_list":
            if user_id != ADMIN_ID: return
            stocks = get_all_stock()
            if not stocks:
                bot.send_message(chat_id, "📦 Stock khali")
                return
            text = "📊 **Stock List**\n\n"
            for cat, prod, count in stocks: text += f"• {cat} | {prod} : {count} pcs\n"
            bot.send_message(chat_id, text, parse_mode="Markdown")
        elif call.data == "admin_pending":
            if user_id != ADMIN_ID: return
            try:
                c.execute("SELECT id, user_id, product, price FROM orders WHERE status='Pending' ORDER BY id DESC")
                orders = c.fetchall()
                if not orders:
                    bot.send_message(chat_id, "✅ No Pending Orders")
                    return
                for o in orders:
                    markup = InlineKeyboardMarkup()
                    markup.add(InlineKeyboardButton("✅ Approve & Send Code", callback_data=f"approve_{o[0]}"))
                    bot.send_message(chat_id, f"🛒 Order ID: {o[0]}\nUser: {o[1]}\nProduct: {o[2]}\nPrice: {o[3]} BDT", reply_markup=markup)
            except Exception as e:
                from database import conn; conn.rollback()
                bot.send_message(chat_id, f"Error: {e}")
        elif call.data.startswith("approve_"):
            if user_id != ADMIN_ID: return
            order_id = int(call.data.split("_")[1])
            bot.send_message(chat_id, f"📦 Enter Product Code for Order {order_id}\n\nVPN: mail:pass\nProxy: host:port:user:pass")
            user_state[user_id] = {"step": "admin_code", "order_id": order_id}
        elif call.data.startswith("confirm_"):
            if user_id != ADMIN_ID: return
            parts = call.data.split("_")
            target_user = int(parts[1]); amount = float(parts[2])
            update_balance(target_user, amount)
            new_balance = get_balance(target_user)
            bonus_to = activate_referral_bonus(target_user, amount)
            if bonus_to:
                try: bot.send_message(bonus_to, f"🎉 Refer Bonus! {target_user} {amount:.0f} BDT deposit korse, tai tumi 0.50 BDT paiso!")
                except: pass
            bot.send_message(target_user, f"✅ ডিপোজিট সফল!\n\n💰 যোগ হয়েছে: +{amount:.2f} টাকা\n💳 ব্যালেন্স: {new_balance:.2f} টাকা")
            try: bot.edit_message_text(f"✅ Confirmed. {amount} BDT added to {target_user}", chat_id=chat_id, message_id=msg_id)
            except: pass
        elif call.data.startswith("cancel_"):
            if user_id != ADMIN_ID: return
            target_user = int(call.data.split("_")[1])
            bot.send_message(target_user, "⚠ Payment Verification Failed")
            try: bot.edit_message_text("❌ Cancelled by Admin", chat_id=chat_id, message_id=msg_id)
            except: pass
        elif call.data == "wallet":
            try: bot.edit_message_text(f"👛 Wallet\n💰 Balance: {get_balance(user_id)} BDT", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "orders":
            orders = get_orders(user_id)
            text = "📦 My Orders\nNo orders yet" if not orders else "📦 My Orders\n"+"\n".join([f"• {x[0]} - {x[1]} BDT - {x[2]}" for x in orders])
            try: bot.edit_message_text(text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "support":
            support_text = "🛡 **SUPPORT UPDATE**\n\n📩 সমস্যা হলে Support: @PolasChandra\n\n🕒 24/7 Assistance"
            try: bot.edit_message_text(support_text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "about":
            try: bot.edit_message_text(f"ℹ About {BOT_NAME}", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "home":
            try: bot.edit_message_text(f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "refer":
            count, earn = get_refer_stats(user_id)
            bot_username = bot.get_me().username
            link = f"https://t.me/{bot_username}?start={user_id}"
            text = f"👥 **Refer & Earn**\n\n🔗 Tomar Link:\n`{link}`\n\n👤 Total Refer: {count}\n💰 Earn: {earn:.2f} BDT"
            try: bot.edit_message_text(text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu(), parse_mode="Markdown")
            except: pass
        try: bot.answer_callback_query(call.id)
        except: pass

    @bot.message_handler(func=lambda m: m.from_user.id in user_state, content_types=['text', 'document'])
    def process_all(message):
        user_id = message.from_user.id
        if is_maintenance_block(user_id):
            bot.reply_to(message, MAINT_MSG)
            return
        if user_id != ADMIN_ID and not is_user_joined(bot, user_id):
            bot.send_message(message.chat.id, f"⚠ Age Channel Join Koro!\n{FORCE_JOIN_CHANNEL}", reply_markup=force_join_menu())
            return
        state = user_state.get(user_id)
        if not state: return
        if state["step"] == "wait_txt_file":
            if message.content_type == 'document' and message.document:
                try:
                    file_name = message.document.file_name.lower()
                    codes = []
                    file_info = bot.get_file(message.document.file_id)
                    downloaded_file = bot.download_file(file_info.file_path)
                    if file_name.endswith(".txt"):
                        codes = downloaded_file.decode("utf-8", errors="ignore").splitlines()
                        codes = [c.strip() for c in codes if c.strip() != ""]
                    elif file_name.endswith(".xlsx"):
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(downloaded_file))
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            if row and row[0]: codes.append(str(row[0]).strip())
                    else: bot.send_message(message.chat.id, "❌ Sudhu.txt ba.xlsx"); return
                    if not codes: bot.send_message(message.chat.id, "❌ File faka!"); return
                    state["codes"] = codes
                    state["step"] = "stock_category"
                    bot.send_message(message.chat.id, f"✅ {len(codes)} ta code peyechi\n\nCategory likho:\n`proxy` / `morelogin` / `edumail` / `hotmail` / `gmail` / `outlook` / `vpn`", parse_mode="Markdown")
                except Exception as e: bot.send_message(message.chat.id, f"❌ Error: {e}")
            else: bot.send_message(message.chat.id, "❌ File upload koro")
        elif state["step"] == "stock_category":
            cat = message.text.lower().strip()
            allowed = ["proxy", "morelogin", "edumail", "hotmail", "gmail", "outlook", "vpn"]
            if cat not in allowed:
                bot.send_message(message.chat.id, "❌ Vul category. `proxy` / `morelogin` / `edumail` / `hotmail` / `gmail` / `outlook` / `vpn` likho", parse_mode="Markdown")
                return
            state["category"] = cat
            state["step"] = "stock_product"
            bot.send_message(message.chat.id, f"✅ Category: {cat}\nEkhon Product er exact name likho (Example: `Morelogin 100 Minutes` / `Outlook.com` / `Edu Gmail Live 10 Minute`)")
        elif state["step"] == "stock_product":
            prod_name = message.text.strip()
            if state["category"] == "morelogin" and "morelogin" in prod_name.lower():
                prod_name = "Morelogin 100 Minutes"
            add_stock(state["category"], prod_name, state["codes"])
            bot.send_message(message.chat.id, f"✅ Stock Add Done!\n{state['category']} | {prod_name} : {len(state['codes'])} pcs")
            del user_state[user_id]
        elif state["step"] == "admin_user_id":
            state["target_id"] = int(message.text); state["step"] = "admin_amount"; bot.send_message(message.chat.id, "💰 Koto BDT?")
        elif state["step"] == "admin_amount":
            amount = float(message.text); target_id = state["target_id"]; update_balance(target_id, amount)
            bot.send_message(message.chat.id, f"✅ {target_id} ke {amount} BDT add"); bot.send_message(target_id, f"🎉 Admin {amount} BDT add korse"); del user_state[user_id]
        elif state["step"] == "broadcast_msg":
            all_users = get_all_users(); sent = 0
            for uid in all_users:
                try: bot.send_message(uid, f"📢 **Notice from {BOT_NAME}**\n\n{message.text}", parse_mode="Markdown"); sent += 1
                except: pass
            bot.send_message(message.chat.id, f"✅ Broadcast Done! {sent} jon ke pathano hoise"); del user_state[user_id]
        elif state["step"] == "admin_code":
            order_id = state["order_id"]
            raw_code = message.text.strip()
            order = get_order_by_id(order_id)
            if not order:
                bot.send_message(message.chat.id, f"❌ Order {order_id} nai")
                del user_state[user_id]
                return
            user_to_send = order[1]
            prod = order[2]
            update_order_status(order_id, "Approved")
            colon_count = raw_code.count(":")
            is_proxy_manual = colon_count >= 3
            is_vpn_manual = (":" in raw_code or "|" in raw_code) and "@" in raw_code.split(":")[0]
            if is_proxy_manual and "vpn" not in prod.lower():
                delivery_text = format_proxy_message(prod, 1, 0, [raw_code], is_manual=True)
                bot.send_message(user_to_send, delivery_text, parse_mode="Markdown")
            elif is_vpn_manual or "vpn" in prod.lower():
                mail = raw_code
                pwd = ""
                if ":" in raw_code:
                    p = raw_code.split(":", 1)
                    mail = p[0].strip()
                    pwd = p[1].strip()
                elif "|" in raw_code:
                    p = raw_code.split("|", 1)
                    mail = p[0].strip()
                    pwd = p[1].strip()
                if pwd:
                    delivery_text = f"✅ DELIVERY SUCCESSFUL!\n───────────────\n\n📧 Mail: `{mail}`\n🔐 Pass: `{pwd}`\n\n\n✅ আপনার অর্ডারটি কমপ্লিট হয়েছে!\nআমাদের উপর ভরসা রাখার জন্য ধন্যবাদ। ❤"
                else:
                    delivery_text = f"✅ DELIVERY SUCCESSFUL!\n───────────────\n\n📧 Mail: `{mail}`\n\n\n✅ আপনার অর্ডারটি কমপ্লিট হয়েছে!\nআমাদের উপর ভরসা রাখার জন্য ধন্যবাদ। ❤"
                bot.send_message(user_to_send, delivery_text, parse_mode="Markdown")
            else:
                bot.send_message(user_to_send, f"✅ Your Order Approved!\n\n📦 Product: {prod}\n🔑 Code:\n`{raw_code}`", parse_mode="Markdown")
            bot.send_message(message.chat.id, f"✅ Order {order_id} Approved & Code sent to {user_to_send}")
            del user_state[user_id]
        elif state["step"] == "amount":
            state["amount"] = message.text; state["step"] = "trx"; bot.send_message(message.chat.id, "🧾 Send Transaction ID / TrxID")
        elif state["step"] == "trx":
            amount = state['amount']; trx = message.text
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton("✅ Confirm", callback_data=f"confirm_{user_id}_{amount}"), InlineKeyboardButton("❌ Cancel", callback_data=f"cancel_{user_id}"))
            bot.send_message(ADMIN_ID,f"💰 New Deposit Request\n👤 {message.from_user.first_name}\n🆔 {user_id}\nAmount: {amount} BDT\nTRX ID: {trx}", reply_markup=markup)
            bot.send_message(message.chat.id,"✅ Deposit Request Sent."); del user_state[user_id]
        elif state["step"] == "custom_qty":
            try:
                qty = int(message.text)
                if qty < 1: qty = 1
                bot.send_message(message.chat.id, f"🛒 {state['name']}\nQuantity: {qty}", reply_markup=quantity_menu(state['category'], state['name'], state['price'], qty))
                del user_state[user_id]
            except: bot.send_message(message.chat.id, "❌ Sothik number dao")

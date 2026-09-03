import io
import os
from datetime import datetime
from database import conn, c, USE_POSTGRES

def create_backup_excel():
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        
        # Users sheet
        ws1 = wb.active
        ws1.title = "Users_Balance"
        ws1.append(["user_id", "balance", "referred_by", "referral_count", "total_referral_earning"])
        try:
            c.execute("SELECT user_id, balance, referred_by, referral_count, total_referral_earning FROM users")
            for row in c.fetchall():
                ws1.append(list(row))
        except Exception as e:
            ws1.append([f"Error: {e}"])

        # Orders sheet
        ws2 = wb.create_sheet("Orders")
        ws2.append(["id", "user_id", "product", "price", "status"])
        try:
            if USE_POSTGRES:
                c.execute("SELECT id, user_id, product, price, status FROM orders ORDER BY id DESC")
            else:
                c.execute("SELECT id, user_id, product, price, status FROM orders ORDER BY id DESC")
            for row in c.fetchall():
                ws2.append(list(row))
        except Exception as e:
            ws2.append([f"Error: {e}"])

        # Stock sheet - AVAILABLE
        ws3 = wb.create_sheet("Stock_Available")
        ws3.append(["id", "category", "product_name", "code", "status"])
        try:
            if USE_POSTGRES:
                c.execute("SELECT id, category, product_name, code, status FROM stock WHERE status='available' ORDER BY id DESC LIMIT 5000")
            else:
                c.execute("SELECT id, category, product_name, code, status FROM stock WHERE status='available' ORDER BY id DESC LIMIT 5000")
            for row in c.fetchall():
                ws3.append(list(row))
        except Exception as e:
            ws3.append([f"Error: {e}"])

        # Stock Used
        ws4 = wb.create_sheet("Stock_Used")
        ws4.append(["id", "category", "product_name", "code", "status"])
        try:
            if USE_POSTGRES:
                c.execute("SELECT id, category, product_name, code, status FROM stock WHERE status='used' ORDER BY id DESC LIMIT 5000")
            else:
                c.execute("SELECT id, category, product_name, code, status FROM stock WHERE status='used' ORDER BY id DESC LIMIT 5000")
            for row in c.fetchall():
                ws4.append(list(row))
        except Exception as e:
            ws4.append([f"Error: {e}"])

        # Referrals
        ws5 = wb.create_sheet("Referrals")
        ws5.append(["id", "referrer_id", "referred_id", "status"])
        try:
            c.execute("SELECT id, referrer_id, referred_id, status FROM referrals")
            for row in c.fetchall():
                ws5.append(list(row))
        except Exception as e:
            ws5.append([f"Error: {e}"])

        # Adjust width
        for ws in [ws1, ws2, ws3, ws4, ws5]:
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        file_stream.name = f"ProxyStore_Backup_{timestamp}.xlsx"
        return file_stream
    except Exception as e:
        print(f"Backup error: {e}")
        return None

def get_db_file_if_sqlite():
    if not USE_POSTGRES and os.path.exists("proxystore.db"):
        return "proxystore.db"
    return None

import os
import openpyxl
from database import conn, c, USE_POSTGRES, create_user

# ========= AUTO RESTORE FORCE =========
try:
    files = [f for f in os.listdir(".") if f.lower().endswith(".xlsx") and "backup" in f.lower()]
    print(f"Found backup files: {files}")
    if files:
        wb = openpyxl.load_workbook(files[0])
        print(f"Sheets: {wb.sheetnames}")
        
        # Check if users table empty
        c.execute("SELECT COUNT(*) FROM users")
        count_before = c.fetchone()[0]
        print(f"Users before restore: {count_before}")
        
        if count_before == 0:  # Only restore if empty
            if "Users_Balance" in wb.sheetnames:
                ws = wb["Users_Balance"]
                restored = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None: continue
                    try:
                        user_id = int(row[0])
                        balance = float(row[1] or 0)
                        referred_by = row[2]
                        ref_count = int(row[3] or 0)
                        ref_earn = float(row[4] or 0)
                        create_user(user_id)
                        if USE_POSTGRES:
                            c.execute("UPDATE users SET balance=%s, referred_by=%s, referral_count=%s, total_referral_earning=%s WHERE user_id=%s",
                                      (balance, referred_by, ref_count, ref_earn, user_id))
                        else:
                            c.execute("UPDATE users SET balance=?, referred_by=?, referral_count=?, total_referral_earning=? WHERE user_id=?",
                                      (balance, referred_by, ref_count, ref_earn, user_id))
                        restored += 1
                    except Exception as e:
                        print(f"Restore row error {row}: {e}")
                conn.commit()
                print(f"✅ RESTORED {restored} USERS")
            
            if "Stock_Available" in wb.sheetnames:
                ws = wb["Stock_Available"]
                restored = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]: continue
                    try:
                        category = str(row[1]).strip()
                        product_name = str(row[2]).strip()
                        code = str(row[3]).strip()
                        if not code: continue
                        if USE_POSTGRES:
                            c.execute("INSERT INTO stock (category, product_name, code, status) VALUES (%s,%s,%s,'available')", (category, product_name, code))
                        else:
                            c.execute("INSERT INTO stock (category, product_name, code, status) VALUES (?,?,?,'available')", (category, product_name, code))
                        restored += 1
                    except Exception as e:
                        print(f"Stock error: {e}")
                conn.commit()
                print(f"✅ RESTORED {restored} STOCK")
        else:
            print("DB not empty, skipping restore")
except Exception as e:
    print(f"Auto restore failed: {e}")
    import traceback
    traceback.print_exc()

# ========= BOT START =========
from bot import bot
from handler import register_handlers
from backup import create_backup_excel, get_db_file_if_sqlite
from config import ADMIN_ID
import threading
import time

register_handlers(bot)

@bot.message_handler(commands=["backup"])
def backup_command(message):
    if message.from_user.id != ADMIN_ID:
        bot.reply_to(message, "❌ Access Denied")
        return
    bot.reply_to(message, "⏳ Backup create hocche...")
    try:
        excel_file = create_backup_excel()
        if excel_file:
            bot.send_document(ADMIN_ID, excel_file, caption=f"FULL BACKUP Time: {excel_file.name}")
        db_file = get_db_file_if_sqlite()
        if db_file:
            with open(db_file, 'rb') as f:
                bot.send_document(ADMIN_ID, f, caption="proxystore.db")
        bot.send_message(ADMIN_ID, "Backup Done!")
    except Exception as e:
        bot.send_message(ADMIN_ID, f"Backup error: {e}")

def auto_backup_loop():
    while True:
        try:
            time.sleep(12 * 60 * 60)
            print("Auto backup running...")
            excel_file = create_backup_excel()
            if excel_file:
                bot.send_document(ADMIN_ID, excel_file, caption=f"AUTO BACKUP 12h - {excel_file.name}")
        except Exception as e:
            print(f"Auto backup error: {e}")
            time.sleep(60)

backup_thread = threading.Thread(target=auto_backup_loop, daemon=True)
backup_thread.start()

print("✅ ProxyStore BOT Started")
bot.infinity_polling(timeout=30, skip_pending=True)
